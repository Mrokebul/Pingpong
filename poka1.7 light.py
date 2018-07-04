######### Edit it by all means :) #################
## writer : Mr. Rokebul ####
### This is the crudest code you can find out there! ##########
#### This thing can ping and spit out the results in an excel file ######

import os, glob, os.path, time
char1='['
char2=']'
char3='('
char4='%'
char5='age ='

####______________________ Functions// Search n Destroy ____________________________###########

def snd_ip (x):
    for line in open(x):
        if "[" in line:
            ip=line[line.find(char1) : line.find(char2)]
            ip=ip[1:]
            open(x).close()
            return ip


def snd_los(x):
    for line in open(x):
        if "(" in line:
            loss_palo=line[line.find(char3):line.find(char4)]
            los=loss_palo[1:]
            open(x).close()
            return los

        
def snd_avg(x):
    for line in open(x):
        if ", Ave" in line:
            Avg_palo=line[line.find(char5):]
            avg=Avg_palo[6:-1]
            open(x).close()
            return avg

####______________________ Functions// Searched n Destroyed ____________________________###########


#-------------------------------------------LAT COMP FUNction------------------------------------------------------------------#
def comp_caller(bex_palo,bex_gmail,bex_micro,bex_apple,bex_itune,bex_cnn,bex_bbc,bex_yout,bex_fb):
    res1 = os.system("ping -n 6 -w 2000 " + bex_palo)
    res2 = os.system("ping -n 6 -w 2000 " + bex_gmail)
    res3 = os.system("ping -n 6 -w 2000 " + bex_micro)
    res4 = os.system("ping -n 6 -w 2000 " + bex_apple)
    res5 = os.system("ping -n 6 -w 2000 " + bex_itune)
    res6 = os.system("ping -n 6 -w 2000 " + bex_cnn)
    res7 = os.system("ping -n 6 -w 2000 " + bex_bbc)
    res8 = os.system("ping -n 6 -w 2000 " + bex_yout)
    res9 = os.system("ping -n 6 -w 2000 " + bex_fb)

    filelist = glob.glob("com*.bat")
    for f in filelist:
        os.remove(f)


    filelist = glob.glob("c*.txt")
    for f in filelist:
        os.remove(f)

    time.sleep(5)
    f = open('Comp_caller.bat','w+')
    f.write('''@echo off

    start cmd /c call com_palo.bat
    start cmd /c call com_gmail.bat
    start cmd /c call com_micro.bat
    start cmd /c call com_apple.bat
    start cmd /c call com_itune.bat
    start cmd /c call com_cnn.bat
    start cmd /c call com_bbc.bat
    start cmd /c call com_youtube.bat
    start cmd /c call com_fb.bat
    ''')
    f.close()

    if res1==0:
        f = open('com_palo.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 5000 '+bex_palo+') > cpalo.txt')
        f.close()

    if res2==0:
        f = open('com_gmail.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_gmail+') > cgmail.txt')
        f.close()

    if res3==0:
        f = open('com_micro.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_micro+') > cmicro.txt')
        f.close()

    if res4==0:
        f = open('com_apple.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_apple+') > capple.txt')
        f.close()

    if res5==0:
        f = open('com_itune.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_itune+') > citune.txt')
        f.close()

    if res6==0:
        f = open('com_cnn.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_cnn+') > ccnn.txt')
        f.close()

    if res7==0:
        f = open('com_bbc.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_bbc+') > cbbc.txt')
        f.close()

    if res8==0:
        f = open('com_youtube.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_yout+') > cyout.txt')
        f.close()

    if res9==0:
        f = open('com_fb.bat','w+')
        f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 '+bex_fb+') > cfb.txt')
        f.close()

    return ("Comp_caller.bat")
#---------------------------------------------------------------------------------------------------------------#


#+++++++++++++++++++++++++++++++++++ BATCH CREATOR++++++++++++++++++++++++++++++++++++++##

#bol-sam
f = open('beximco.bat','w+')
f.write('''@echo on
netsh wlan disconnect
timeout 5
netsh wlan connect BOL-SAM
timeout 5
netsh wlan connect BOL-SAM
timeout 5
''')
f.close()

#aa
f = open('aamra.bat','w+')
f.write('''@echo on
netsh wlan disconnect
timeout 5
netsh wlan connect AA
timeout 5
netsh wlan connect AA
timeout 5
''')
f.close()

#LN
f = open('link3.bat','w+')
f.write('''@echo on
netsh wlan disconnect
timeout 5
netsh wlan connect LN
timeout 5
netsh wlan connect LN
timeout 5
''')
f.close()

#ag
f = open('agni.bat','w+')
f.write('''@echo on
netsh wlan disconnect
timeout 5
netsh wlan connect AG
timeout 5
netsh wlan connect AG
timeout 5
''')
f.close()

#dze
f = open('doze.bat','w+')
f.write('''@echo on
netsh wlan disconnect
timeout 5
netsh wlan connect DZE
timeout 5
netsh wlan connect DZE
timeout 5
''')
f.close()

######################################3
f = open('caller.bat','w+')
f.write('''@echo off

start cmd /c call palo.bat
start cmd /c call gmail.bat
start cmd /c call micro.bat
start cmd /c call apple.bat
start cmd /c call itune.bat
start cmd /c call cnn.bat
start cmd /c call bbc.bat
start cmd /c call youtube.bat
start cmd /c call fb.bat
''')
f.close()

f = open('palo.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 prothom-alo.com) > palo.txt')
f.close()

f = open('gmail.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 mail.google.com) > gmail.txt')
f.close()

f = open('micro.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 download.microsoft.com) > micro.txt')
f.close()

f = open('apple.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 www.apple.com) > apple.txt')
f.close()

f = open('itune.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 itunes.apple.com) > itune.txt')
f.close()

f = open('cnn.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 www.cnn.com) > cnn.txt')
f.close()

f = open('bbc.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 www.bbc.co.uk) > bbc.txt')
f.close()

f = open('youtube.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 www.youtube.com) > yout.txt')
f.close()

f = open('fb.bat','w+')
f.write('@echo on '+'\n'+'(ping -n 200 -w 6000 www.facebook.com) > fb.txt')
f.close()


##########----------------+++++++ END OF BAT CREATION++++++++----------------###############

from openpyxl import load_workbook

wb = load_workbook('template.xlsx')

# grab the active worksheet
ws = wb.active
sheet=wb.active

#################connect beximco  #################################################################################################

os.system("beximco.bat")
time.sleep(4)
################### check if the BEXIMCO is up ###############
response = os.system("ping -n 5 8.8.8.8")

#and then check the response...
if response == 0:
    
    os.system("caller.bat")
    time.sleep(250)


############### prothom alo ##############################
    bex_palo=snd_ip('palo.txt')
    los1=snd_los('palo.txt')
    avg1= snd_avg('palo.txt')
###################### gmail ###############################
    bex_gmail=snd_ip('gmail.txt')
    los2=snd_los('gmail.txt')
    avg2= snd_avg('gmail.txt')
######################### microsoft ################################
    bex_micro=snd_ip('micro.txt')
    los3=snd_los('micro.txt')
    avg3= snd_avg('micro.txt')
######################### apple ################################
    bex_apple=snd_ip('apple.txt')
    los4=snd_los('apple.txt')
    avg4= snd_avg('apple.txt')
######################### itune ################################
    bex_itune=snd_ip('itune.txt')
    los5=snd_los('itune.txt')
    avg5= snd_avg('itune.txt')
######################### cnn ################################
    bex_cnn=snd_ip('cnn.txt')
    los6=snd_los('cnn.txt')
    avg6= snd_avg('cnn.txt')
######################### bbc ################################
    bex_bbc=snd_ip('bbc.txt')
    los7=snd_los('bbc.txt')
    avg7= snd_avg('bbc.txt')
######################### youtube ################################
    bex_yout=snd_ip('yout.txt')
    los8=snd_los('yout.txt')
    avg8= snd_avg('yout.txt')
######################### facebook ################################
    bex_fb=snd_ip('fb.txt')
    los9=snd_los('fb.txt')
    avg9= snd_avg('fb.txt')

#############         paster   #####################################################
    
    IP=[bex_palo,bex_gmail,bex_micro,bex_apple,bex_itune,bex_cnn,bex_bbc,bex_yout,bex_fb]
    LOS=[los1,los2,los3,los4,los5,los6,los7,los8,los9]
    AVG=[avg1,avg2,avg3,avg4,avg5,avg6,avg7,avg8,avg9]

    x=0
    y=0
    z=0
    
    for a in range (3,6):
        for b in range(4,13):
            if a==5:   #data will be only in col 3
                sheet.cell(row=b, column=a).value=IP[x]
                x=x+1
            elif a==4:
                sheet.cell(row=b, column=a).value=LOS[y]
                y=y+1
            else:
                sheet.cell(row=b, column=a).value=AVG[z]
                z=z+1
    x=0
    y=0
    z=0
    for a in range (3,6):
        for b in range(18,27):
            if a==3:   #data will be only in col 3
                sheet.cell(row=b, column=a).value=IP[x]
                x=x+1
            elif a==4:
                sheet.cell(row=b, column=a).value=LOS[y]
                y=y+1
            else:
                sheet.cell(row=b, column=a).value=AVG[z]
                z=z+1
                
###############################################################

else:
    print ('Beximco down!')
    
    ######### link down in excel ###########

    for i in range(4,13):
        for c in range(3,6):
            sheet.cell(row=i, column=c).value='N/A'



#################    connect Link3    ##################################################################################################

os.system("link3.bat")
time.sleep(4)
print ('Link3 connected!')
################### check if the Link3 is up ###############
response = os.system("ping -n 5 8.8.8.8")

#and then check the response...
if response == 0:
    comp_caller(bex_palo,bex_gmail,bex_micro,bex_apple,bex_itune,bex_cnn,bex_bbc,bex_yout,bex_fb)

    os.system("caller.bat")
        
    os.system('Comp_caller.bat')

    time.sleep(250)


############### prothom alo ##############################


    ws['K4'] = snd_ip('palo.txt')

    ws['J4'] = snd_los('palo.txt')

    ws['I4'] = snd_avg('palo.txt')

###################### gmail ###############################

    ws['K5'] = snd_ip('gmail.txt')

    ws['J5'] = snd_los('gmail.txt')

    ws['I5'] = snd_avg('gmail.txt')

######################### microsoft ################################


    ws['K6'] = snd_ip('micro.txt')

    ws['J6'] = snd_los('micro.txt')

    ws['I6'] = snd_avg('micro.txt')

######################### apple ################################

    ws['K7'] = snd_ip('apple.txt')

    ws['J7'] = snd_los('apple.txt')

    ws['I7'] = snd_avg('apple.txt')


######################### itune ################################

    ws['K8'] = snd_ip('itune.txt')

    ws['J8'] = snd_los('itune.txt')

    ws['I8'] = snd_avg('itune.txt')

######################### cnn ################################


    ws['K9'] = snd_ip('cnn.txt')

    ws['J9'] = snd_los('cnn.txt')

    ws['I9'] = snd_avg('cnn.txt')

######################### bbc ################################


    ws['K10'] = snd_ip('bbc.txt')

    ws['J10'] = snd_los('bbc.txt')

    ws['I10'] = snd_avg('bbc.txt')

######################### youtube ################################


    ws['K11'] = snd_ip('yout.txt')

    ws['J11'] = snd_los('yout.txt')

    ws['I11'] = snd_avg('yout.txt')

######################### facebook ################################


    ws['K12'] = snd_ip('fb.txt')

    ws['J12'] = snd_los('fb.txt')

    ws['I12'] = snd_avg('fb.txt')



  
########################         LATENCY COMPARISON TEST         #######################################


############### prothom alo ##############################

    a=glob.glob("cpalo.txt")
    if a:
        ws['I18'] = snd_los('cpalo.txt')
        ws['H18'] = snd_avg('cpalo.txt')

###################### gmail ###############################
    b=glob.glob("cgmail.txt")
    if b:
        ws['I19'] = snd_los('cgmail.txt')
        ws['H19'] = snd_avg('cgmail.txt')

######################### microsoft ################################

    c=glob.glob("cmicro.txt")
    if c:
        ws['I20'] = snd_los('cmicro.txt')
        ws['H20'] = snd_avg('cmicro.txt')

######################### apple ################################

    d=glob.glob("capple.txt")
    if d:
        ws['I21'] = snd_los('capple.txt')
        ws['H21'] = snd_avg('capple.txt')


######################### itune ################################

    e=glob.glob("citune.txt")
    if e:
        ws['I22'] = snd_los('citune.txt')
        ws['H22'] = snd_avg('citune.txt')

######################### cnn ################################

    f=glob.glob("ccnn.txt")
    if f:
        ws['I23'] = snd_los('ccnn.txt')
        ws['H23'] = snd_avg('ccnn.txt')


######################### bbc ################################

    g=glob.glob("cbbc.txt")
    if g:
        ws['I24'] = snd_los('cbbc.txt')
        ws['H24'] = snd_avg('cbbc.txt')


######################### youtube ################################

    h=glob.glob("cyout.txt")
    if h:
        ws['I25'] = snd_los('cyout.txt')
        ws['H25'] = snd_avg('cyout.txt')

######################### facebook ################################

    i=glob.glob("cfb.txt")
    if i:
        ws['I26'] = snd_los('cfb.txt')
        ws['H26'] = snd_avg('cfb.txt')


else:
    print ('Link3 down!')
  ########### link down in xl ############
    for i in range(4,13):
        for c in range(9,12):
            sheet.cell(row=i, column=c).value='N/A'










#################connect Aamra ###################################################################################################

os.system("aamra.bat")
time.sleep(4)
print ('aamra connected')
################### check if the aamra is up ###############
response = os.system("ping -n 5 8.8.8.8")

#and then check the response...
if response == 0:
    comp_caller(bex_palo,bex_gmail,bex_micro,bex_apple,bex_itune,bex_cnn,bex_bbc,bex_yout,bex_fb)
    
    os.system("caller.bat")

    os.system('Comp_caller.bat')
   
    time.sleep(250)


############### prothom alo ##############################

    ws['H4'] = snd_ip('palo.txt')

    ws['G4'] = snd_los('palo.txt')

    ws['F4'] = snd_avg('palo.txt')

###################### gmail ###############################

    ws['H5'] = snd_ip('gmail.txt')

    ws['G5'] = snd_los('gmail.txt')

    ws['F5'] = snd_avg('gmail.txt')

######################### microsoft ################################


    ws['H6'] = snd_ip('micro.txt')

    ws['G6'] = snd_los('micro.txt')

    ws['F6'] = snd_avg('micro.txt')

######################### apple ################################


    ws['H7'] = snd_ip('apple.txt')

    ws['G7'] = snd_los('apple.txt')

    ws['F7'] = snd_avg('apple.txt')


######################### itune ################################

    ws['H8'] = snd_ip('itune.txt')

    ws['G8'] = snd_los('itune.txt')

    ws['F8'] = snd_avg('itune.txt')

######################### cnn ################################


    ws['H9'] = snd_ip('cnn.txt')

    ws['G9'] = snd_los('cnn.txt')

    ws['F9'] = snd_avg('cnn.txt')

######################### bbc ################################

    ws['H10'] = snd_ip('bbc.txt')

    ws['G10'] = snd_los('bbc.txt')

    ws['F10'] = snd_avg('bbc.txt')

######################### youtube ################################

    ws['H11'] = snd_ip('yout.txt')

    ws['G11'] = snd_los('yout.txt')

    ws['F11'] = snd_avg('yout.txt')

######################### facebook ################################


    ws['H12'] = snd_ip('fb.txt')

    ws['G12'] = snd_los('fb.txt')

    ws['F12'] = snd_avg('fb.txt')


     
########################         LATENCY COMPARISON TEST         #######################################


############### prothom alo ##############################

    a=glob.glob("cpalo.txt")
    if a:
        ws['G18'] = snd_los('cpalo.txt')
        ws['F18'] = snd_avg('cpalo.txt')

###################### gmail ###############################

    b=glob.glob("cgmail.txt")
    if b:
        ws['G19'] = snd_los('cgmail.txt')
        ws['F19'] = snd_avg('cgmail.txt')

######################### microsoft ################################

    c=glob.glob("cmicro.txt")
    if c:
        ws['G20'] = snd_los('cmicro.txt')
        ws['F20'] = snd_avg('cmicro.txt')

######################### apple ################################

    d=glob.glob("capple.txt")
    if d:
        ws['G21'] = snd_los('capple.txt')
        ws['F21'] = snd_avg('capple.txt')


######################### itune ################################

    e=glob.glob("citune.txt")
    if e:
        ws['G22'] = snd_los('citune.txt')
        ws['F22'] = snd_avg('citune.txt')

######################### cnn ################################

    f=glob.glob("ccnn.txt")
    if f:
        ws['G23'] = snd_los('ccnn.txt')
        ws['F23'] = snd_avg('ccnn.txt')

######################### bbc ################################

    g=glob.glob("cbbc.txt")
    if g:
        ws['G24'] = snd_los('cbbc.txt')
        ws['F24'] = snd_avg('cbbc.txt')

######################### youtube ################################

    h=glob.glob("cyout.txt")
    if h:
        ws['G25'] = snd_los('cyout.txt')
        ws['F25'] = snd_avg('cyout.txt')

######################### facebook ################################
    i=glob.glob("cfb.txt")
    if i:
        ws['G26'] = snd_los('cfb.txt')
        ws['F26'] = snd_avg('cfb.txt')


else:
    print ('Aamra down!')
    
    ##############link down in xl############
    
    for i in range(4,13):
        for c in range(6,9):
            sheet.cell(row=i, column=c).value='N/A'






#################connect Agni ###########################################################################################################

os.system("agni.bat")
time.sleep(4)
print ('Agni connected!')
################### check if the agni is up ###############
response = os.system("ping -n 5 8.8.8.8")

#and then check the response...
if response == 0:
    comp_caller(bex_palo,bex_gmail,bex_micro,bex_apple,bex_itune,bex_cnn,bex_bbc,bex_yout,bex_fb)

    os.system("caller.bat")

    os.system('Comp_caller.bat')

    time.sleep(250)


############### prothom alo ##############################

    ws['N4'] = snd_ip('palo.txt')


    ws['M4'] = snd_los('palo.txt')

    ws['L4'] = snd_avg('palo.txt')

###################### gmail ###############################


    ws['N5'] = snd_ip('gmail.txt')

    ws['M5'] = snd_los('gmail.txt')

    ws['L5'] = snd_avg('gmail.txt')

######################### microsoft ################################

    ws['N6'] = snd_ip('micro.txt')

    ws['M6'] = snd_los('micro.txt')

    ws['L6'] = snd_avg('micro.txt')

######################### apple ################################


    ws['N7'] = snd_ip('apple.txt')

    ws['M7'] = snd_los('apple.txt')

    ws['L7'] = snd_avg('apple.txt')


######################### itune ################################


    ws['N8'] = snd_ip('itune.txt')

    ws['M8'] = snd_los('itune.txt')

    ws['L8'] = snd_avg('itune.txt')

######################### cnn ################################


    ws['N9'] = snd_ip('cnn.txt')

    ws['M9'] = snd_los('cnn.txt')

    ws['L9'] = snd_avg('cnn.txt')

######################### bbc ################################


    ws['N10'] = snd_ip('bbc.txt')

    ws['M10'] = snd_los('bbc.txt')

    ws['L10'] = snd_avg('bbc.txt')

######################### youtube ################################


    ws['N11'] = snd_ip('yout.txt')

    ws['M11'] = snd_los('yout.txt')

    ws['L11'] = snd_avg('yout.txt')

######################### facebook ################################


    ws['N12'] = snd_ip('fb.txt')

    ws['M12'] = snd_los('fb.txt')

    ws['L12'] = snd_avg('fb.txt')





  ########################         LATENCY COMPARISON TEST         #######################################


    ############### prothom alo ##############################

    a=glob.glob("cpalo.txt")
    if a:
        ws['K18'] = snd_los('cpalo.txt')
        ws['J18'] = snd_avg('cpalo.txt')

    ###################### gmail ###############################
    b=glob.glob("cgmail.txt")
    if b:
        ws['K19'] = snd_los('cgmail.txt')
        ws['J19'] = snd_avg('cgmail.txt')

    ######################### microsoft ################################

    c=glob.glob("cmicro.txt")
    if c:
        ws['K20'] = snd_los('cmicro.txt')
        ws['J20'] = snd_avg('cmicro.txt')

    ######################### apple ################################

    d=glob.glob("capple.txt")
    if d:
        ws['K21'] = snd_los('capple.txt')
        ws['J21'] = snd_avg('capple.txt')

    ######################### itune ################################

    e=glob.glob("citune.txt")
    if e:
        ws['K22'] = snd_los('citune.txt')
        ws['J22'] = snd_avg('citune.txt')

    ######################### cnn ################################

    f=glob.glob("ccnn.txt")
    if f:
        ws['K23'] = snd_los('ccnn.txt')
        ws['J23'] = snd_avg('ccnn.txt')

    ######################### bbc ################################

    g=glob.glob("cbbc.txt")
    if g:
        ws['K24'] = snd_los('cbbc.txt')
        ws['J24'] = snd_avg('cbbc.txt')

    ######################### youtube ################################

    h=glob.glob("cyout.txt")
    if h:
        ws['K25'] = snd_los('cyout.txt')
        ws['J25'] = snd_avg('cyout.txt')

    ######################### facebook ################################

    i=glob.glob("cfb.txt")
    if i:
        ws['K26'] = snd_los('cfb.txt')
        ws['J26'] = snd_avg('cfb.txt')   



else:
    print ('Agni down!')
    for i in range(4,13):
        for c in range(12,15):
            sheet.cell(row=i, column=c).value='N/A'





#################connect Doze ###############################################################################################################

os.system("doze.bat")
time.sleep(4)
print ('doze connected!')
################### check if the Link3 is up ###############
response = os.system("ping -n 6 8.8.8.8")

#and then check the response...
if response == 0:
    comp_caller(bex_palo,bex_gmail,bex_micro,bex_apple,bex_itune,bex_cnn,bex_bbc,bex_yout,bex_fb)

    os.system("caller.bat")

    os.system('Comp_caller.bat')
    
    time.sleep(300)
    ############### prothom alo ##############################


    ws['Q4'] = snd_ip('palo.txt')


    ws['P4'] = snd_los('palo.txt')

    ws['O4'] = snd_avg('palo.txt')

    ###################### gmail ###############################


    ws['Q5'] = snd_ip('gmail.txt')

    ws['P5'] = snd_los('gmail.txt')

    ws['O5'] = snd_avg('gmail.txt')

    ######################### microsoft ################################


    ws['Q6'] = snd_ip('micro.txt')

    ws['P6'] = snd_los('micro.txt')

    ws['O6'] = snd_avg('micro.txt')

    ######################### apple ################################


    ws['Q7'] = snd_ip('apple.txt')

    ws['P7'] = snd_los('apple.txt')

    ws['O7'] = snd_avg('apple.txt')


    ######################### itune ################################


    ws['Q8'] = snd_ip('itune.txt')

    ws['P8'] = snd_los('itune.txt')

    ws['O8'] = snd_avg('itune.txt')

    ######################### cnn ################################


    ws['Q9'] = snd_ip('cnn.txt')

    ws['P9'] = snd_los('cnn.txt')

    ws['O9'] = snd_avg('cnn.txt')

    ######################### bbc ################################


    ws['Q10'] = snd_ip('bbc.txt')

    ws['P10'] = snd_los('bbc.txt')

    ws['O10'] = snd_avg('bbc.txt')

    ######################### youtube ################################


    ws['Q11'] = snd_ip('yout.txt')

    ws['P11'] = snd_los('yout.txt')

    ws['O11'] = snd_avg('yout.txt')

    ######################### facebook ################################


    ws['Q12'] = snd_ip('fb.txt')

    ws['P12'] = snd_los('fb.txt')

    ws['O12'] = snd_avg('fb.txt')



     

  ########################         LATENCY COMPARISON TEST         #######################################

    ############### prothom alo ##############################

    a=glob.glob("cpalo.txt")
    if a:
        ws['M18'] = snd_los('cpalo.txt')
        ws['L18'] = snd_avg('cpalo.txt')

    ###################### gmail ###############################

    b=glob.glob("cgmail.txt")
    if b:
        ws['M19'] = snd_los('cgmail.txt')
        ws['L19'] = snd_avg('cgmail.txt')

    ######################### microsoft ################################

    c=glob.glob("cmicro.txt")
    if c:
        ws['M20'] = snd_los('cmicro.txt')
        ws['L20'] = snd_avg('cmicro.txt')

    ######################### apple ################################

    d=glob.glob("capple.txt")
    if d:
        ws['M21'] = snd_los('capple.txt')
        ws['L21'] = snd_avg('capple.txt')


    ######################### itune ################################

    e=glob.glob("citune.txt")
    if e:
        ws['M22'] = snd_los('citune.txt')
        ws['L22'] = snd_avg('citune.txt')

    ######################### cnn ################################

    f=glob.glob("ccnn.txt")
    if f:
        ws['M23'] = snd_los('ccnn.txt')
        ws['L23'] = snd_avg('ccnn.txt')

    ######################### bbc ################################

    g=glob.glob("cbbc.txt")
    if g:
        ws['M24'] = snd_los('cbbc.txt')
        ws['L24'] = snd_avg('cbbc.txt')

    ######################### youtube ################################

    h=glob.glob("cyout.txt")
    if h:
        ws['M25'] = snd_los('cyout.txt')
        ws['L25'] = snd_avg('cyout.txt')

    ######################### facebook ################################

    i=glob.glob("cfb.txt")
    if i:
        ws['M26'] = snd_los('cfb.txt')
        ws['L26'] = snd_avg('cfb.txt')


        
else:
    print ('Doze down!')
    ########### link down in xl ############
    for i in range(4,13):
        for c in range(15,18):
            sheet.cell(row=i, column=c).value='N/A'




# Save the xl file
wb.save("~~~~Final.xlsx")

filelist = glob.glob("*.bat")
for f in filelist:
    os.remove(f)


filelist = glob.glob("*.txt")
for f in filelist:
    os.remove(f)



